"""
Product Master Routes — Bridge Proxy
─────────────────────────────────────
Proxies all product master requests (Shares, MF, ETF, Life/Health Insurance)
to the tenant's Bridge. Also serves Excel demo templates generated here.
"""
import io
from typing import List

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse

from app.api.deps import get_bridge_client, get_current_user
from app.services.bridge_client import BridgeClient

router = APIRouter()

PRODUCT_TYPES = ("shares", "mutual-funds", "etfs", "life-insurance", "health-insurance")
PRICE_TYPES   = ("share-prices", "nav-uploads", "etf-prices")

EXCEL_COLUMNS = {
    "shares":           ["ISIN Code", "Symbol", "Share Name"],
    "mutual-funds":     ["Scheme Code", "Fund House Name", "Scheme Name", "Asset Category"],
    "etfs":             ["ISIN Code", "Symbol", "ETF Name"],
    "life-insurance":   ["Company Name", "Policy Name", "Policy Type", "UIN"],
    "health-insurance": ["Company Name", "Policy Name", "UIN"],
}

EXCEL_SAMPLES = {
    "shares":           [("INE002A01018", "RELIANCE", "Reliance Industries Limited")],
    "mutual-funds":     [("120503", "HDFC AMC", "HDFC Mid-Cap Opportunities Fund", "Equity")],
    "etfs":             [("INF204KB13I2", "NIFTYBEES", "Nippon India ETF Nifty BeES")],
    "life-insurance":   [("LIC OF INDIA", "JEEVAN ANAND", "Savings Insurance", "101L048V03")],
    "health-insurance": [("STAR HEALTH", "FAMILY HEALTH OPTIMA", "SHAHLGP23001V012223")],
}


PRICE_EXCEL_COLUMNS = {
    "share-prices": ["ISIN Code", "Symbol", "Date (DD-MM-YYYY)", "Share Price"],
    "nav-uploads":  ["Scheme Code", "Scheme Name", "Date (DD-MM-YYYY)", "NAV"],
    "etf-prices":   ["ISIN Code", "Symbol", "Date (DD-MM-YYYY)", "ETF Price"],
}

PRICE_EXCEL_SAMPLES = {
    "share-prices": [("INE002A01018", "RELIANCE", "01-01-2025", "2850.50")],
    "nav-uploads":  [("120503", "Hdfc Mid-Cap Opportunities Fund", "01-01-2025", "145.23")],
    "etf-prices":   [("INF204KB13I2", "NIFTYBEES", "01-01-2025", "248.75")],
}


def _validate_product_type(product_type: str):
    if product_type not in PRODUCT_TYPES:
        raise HTTPException(400, f"Invalid product type '{product_type}'. Must be one of: {', '.join(PRODUCT_TYPES)}")


def _validate_price_type(price_type: str):
    if price_type not in PRICE_TYPES:
        raise HTTPException(400, f"Invalid price type '{price_type}'. Must be one of: {', '.join(PRICE_TYPES)}")


# ── List ────────────────────────────────────────────────────────────

@router.get("/products/{product_type}")
async def list_products(
    product_type: str,
    skip: int = Query(0),
    limit: int = Query(200),
    search: str = Query(None),
    bridge: BridgeClient = Depends(get_bridge_client),
    current_user=Depends(get_current_user),
):
    _validate_product_type(product_type)
    params = {"skip": skip, "limit": limit}
    if search:
        params["search"] = search
    return await bridge.get(f"/products/{product_type}", params=params)


# ── Create ──────────────────────────────────────────────────────────

@router.post("/products/{product_type}")
async def create_product(
    product_type: str,
    data: dict,
    bridge: BridgeClient = Depends(get_bridge_client),
    current_user=Depends(get_current_user),
):
    _validate_product_type(product_type)
    return await bridge.post(f"/products/{product_type}", data)


# ── Update ──────────────────────────────────────────────────────────

@router.patch("/products/{product_type}/{product_id}")
async def update_product(
    product_type: str,
    product_id: str,
    data: dict,
    bridge: BridgeClient = Depends(get_bridge_client),
    current_user=Depends(get_current_user),
):
    _validate_product_type(product_type)
    return await bridge.patch(f"/products/{product_type}/{product_id}", data)


# ── Toggle Active/Inactive ──────────────────────────────────────────

@router.patch("/products/{product_type}/{product_id}/toggle")
async def toggle_product(
    product_type: str,
    product_id: str,
    bridge: BridgeClient = Depends(get_bridge_client),
    current_user=Depends(get_current_user),
):
    _validate_product_type(product_type)
    return await bridge.patch(f"/products/{product_type}/{product_id}/toggle")


# ── Research Reports: List ──────────────────────────────────────────

@router.get("/products/{product_type}/{product_id}/reports")
async def list_reports(
    product_type: str,
    product_id: str,
    bridge: BridgeClient = Depends(get_bridge_client),
    current_user=Depends(get_current_user),
):
    _validate_product_type(product_type)
    return await bridge.get(f"/products/{product_type}/{product_id}/reports")


# ── Research Reports: Upload ────────────────────────────────────────

@router.post("/products/{product_type}/{product_id}/reports")
async def upload_reports(
    product_type: str,
    product_id: str,
    files: List[UploadFile] = File(...),
    uploader_name: str = Form(None),
    bridge: BridgeClient = Depends(get_bridge_client),
    current_user=Depends(get_current_user),
):
    _validate_product_type(product_type)

    file_tuples = []
    for f in files:
        content = await f.read()
        file_tuples.append(("files", (f.filename, content, f.content_type or "application/octet-stream")))

    form_data = {}
    if uploader_name:
        form_data["uploader_name"] = uploader_name
    elif hasattr(current_user, "email"):
        form_data["uploader_name"] = getattr(current_user, "email", "")

    return await bridge.post(
        f"/products/{product_type}/{product_id}/reports",
        data=form_data,
        files=file_tuples,
    )


# ── Research Reports: Download ──────────────────────────────────────

@router.get("/products/reports/{report_id}/download")
async def download_report(
    report_id: str,
    bridge: BridgeClient = Depends(get_bridge_client),
    current_user=Depends(get_current_user),
):
    return await bridge.get(f"/products/reports/{report_id}/download")


# ── Excel Preview ───────────────────────────────────────────────────

@router.post("/products/{product_type}/excel-preview")
async def excel_preview(
    product_type: str,
    file: UploadFile = File(...),
    bridge: BridgeClient = Depends(get_bridge_client),
    current_user=Depends(get_current_user),
):
    _validate_product_type(product_type)
    content = await file.read()
    files = {"file": (file.filename, content, file.content_type or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    return await bridge.post(f"/products/{product_type}/excel-preview", files=files)


# ── Excel Import (confirm) ──────────────────────────────────────────

@router.post("/products/{product_type}/excel-import")
async def excel_import(
    product_type: str,
    payload: dict,
    bridge: BridgeClient = Depends(get_bridge_client),
    current_user=Depends(get_current_user),
):
    _validate_product_type(product_type)
    return await bridge.post(f"/products/{product_type}/excel-import", payload)


# ── Excel Template Download ─────────────────────────────────────────

@router.get("/products/{product_type}/excel-template")
def download_excel_template(
    product_type: str,
    current_user=Depends(get_current_user),
):
    """Generate and return a demo Excel template for the given product type."""
    _validate_product_type(product_type)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl is not installed on the backend server.")

    columns = EXCEL_COLUMNS[product_type]
    sample_rows = EXCEL_SAMPLES.get(product_type, [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = product_type.replace("-", " ").title()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = max(20, len(col_name) + 4)

    for row_idx, sample in enumerate(sample_rows, start=2):
        for col_idx, val in enumerate(sample, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{product_type}_template.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Price Upload: List ──────────────────────────────────────────────

@router.get("/price-uploads/{price_type}")
async def list_price_records(
    price_type: str,
    skip: int = Query(0),
    limit: int = Query(200),
    search: str = Query(None),
    bridge: BridgeClient = Depends(get_bridge_client),
    current_user=Depends(get_current_user),
):
    _validate_price_type(price_type)
    params = {"skip": skip, "limit": limit}
    if search:
        params["search"] = search
    return await bridge.get(f"/price-uploads/{price_type}", params=params)


# ── Price Upload: Create ────────────────────────────────────────────

@router.post("/price-uploads/{price_type}")
async def create_price_record(
    price_type: str,
    data: dict,
    bridge: BridgeClient = Depends(get_bridge_client),
    current_user=Depends(get_current_user),
):
    _validate_price_type(price_type)
    return await bridge.post(f"/price-uploads/{price_type}", data)


# ── Price Upload: Toggle ────────────────────────────────────────────

@router.patch("/price-uploads/{price_type}/{record_id}/toggle")
async def toggle_price_record(
    price_type: str,
    record_id: str,
    bridge: BridgeClient = Depends(get_bridge_client),
    current_user=Depends(get_current_user),
):
    _validate_price_type(price_type)
    return await bridge.patch(f"/price-uploads/{price_type}/{record_id}/toggle")


# ── Price Upload: Excel Preview ─────────────────────────────────────

@router.post("/price-uploads/{price_type}/excel-preview")
async def price_excel_preview(
    price_type: str,
    file: UploadFile = File(...),
    bridge: BridgeClient = Depends(get_bridge_client),
    current_user=Depends(get_current_user),
):
    _validate_price_type(price_type)
    content = await file.read()
    files = {"file": (file.filename, content, file.content_type or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    return await bridge.post(f"/price-uploads/{price_type}/excel-preview", files=files)


# ── Price Upload: Excel Import ──────────────────────────────────────

@router.post("/price-uploads/{price_type}/excel-import")
async def price_excel_import(
    price_type: str,
    payload: dict,
    bridge: BridgeClient = Depends(get_bridge_client),
    current_user=Depends(get_current_user),
):
    _validate_price_type(price_type)
    return await bridge.post(f"/price-uploads/{price_type}/excel-import", payload)


# ── Price Upload: Excel Template ────────────────────────────────────

@router.get("/price-uploads/{price_type}/excel-template")
def download_price_excel_template(
    price_type: str,
    current_user=Depends(get_current_user),
):
    _validate_price_type(price_type)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl is not installed on the backend server.")

    columns = PRICE_EXCEL_COLUMNS[price_type]
    sample_rows = PRICE_EXCEL_SAMPLES.get(price_type, [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = price_type.replace("-", " ").title()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = max(20, len(col_name) + 4)

    for row_idx, sample in enumerate(sample_rows, start=2):
        for col_idx, val in enumerate(sample, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{price_type}_template.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
